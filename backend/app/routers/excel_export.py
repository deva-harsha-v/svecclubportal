import io
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..database import get_db
from ..models import Club, Student, Registration, StaffProfile, Role, AuditLog
from ..auth import require_admin_or_club_head

router = APIRouter(prefix="/api/admin/export", tags=["export"])


def _visible_club_ids(staff: StaffProfile, db: Session) -> list[int] | None:
    if staff.role == Role.ADMIN:
        return None
    if not staff.club_id:
        raise HTTPException(status_code=403, detail="This account is not linked to a club.")
    return [staff.club_id]


def _log_audit(db: Session, staff: StaffProfile, action: str, details: str, ip: str | None = None):
    audit = AuditLog(
        user_id=staff.user_id,
        user_email=staff.email,
        action=action,
        details=details,
        ip_address=ip,
    )
    db.add(audit)
    db.commit()


@router.get("/excel")
def export_excel(
    request: Request,
    club: str | None = Query(None, description="Optional: export only this club (slug)"),
    staff: StaffProfile = Depends(require_admin_or_club_head),
    db: Session = Depends(get_db),
):
    allowed = _visible_club_ids(staff, db)

    q = (
        db.query(Registration)
        .options(joinedload(Registration.student), joinedload(Registration.club))
        .join(Club)
    )

    if allowed is not None:
        q = q.filter(Registration.club_id.in_(allowed))

    if club:
        club_obj = db.query(Club).filter(Club.slug == club.lower()).first()
        if not club_obj:
            raise HTTPException(status_code=404, detail="Club not found.")
        if allowed is not None and club_obj.id not in allowed:
            raise HTTPException(status_code=403, detail="You cannot export another club's data.")
        q = q.filter(Registration.club_id == club_obj.id)

    regs = q.order_by(Club.name, Registration.registered_at).all()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="12193B", end_color="12193B", fill_type="solid")
    title_font = Font(name="Arial", size=14, bold=True, color="12193B")

    # 1. SUMMARY SHEET
    if staff.role == Role.ADMIN and not club:
        summary_ws = wb.create_sheet(title="Summary")
        summary_ws.views.sheetView[0].showGridLines = True

        summary_ws.append(["Club Discovery & Registration Portal — Summary Report"])
        summary_ws.cell(row=1, column=1).font = title_font
        summary_ws.append([f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}"])
        summary_ws.append([])

        total_students = db.query(func.count(Student.id)).scalar() or 0
        total_regs = db.query(func.count(Registration.id)).scalar() or 0
        total_clubs = db.query(func.count(Club.id)).filter(Club.is_active == True).scalar() or 0

        summary_ws.append(["Metric", "Value"])
        summary_ws.append(["Total Registered Students", total_students])
        summary_ws.append(["Total Club Registrations", total_regs])
        summary_ws.append(["Total Active Clubs", total_clubs])

        for cell in summary_ws[4]:
            cell.font = header_font
            cell.fill = header_fill

        summary_ws.append([])
        summary_ws.append(["Club Name", "Category", "Registrations"])
        for cell in summary_ws[9]:
            cell.font = header_font
            cell.fill = header_fill

        club_counts = (
            db.query(Club.name, Club.category, func.count(Registration.id))
            .outerjoin(Registration, Registration.club_id == Club.id)
            .filter(Club.is_active == True)
            .group_by(Club.id)
            .order_by(func.count(Registration.id).desc())
            .all()
        )
        for cname, ccat, count in club_counts:
            summary_ws.append([cname, ccat, count])

    # 2. ALL REGISTRATIONS SHEET
    all_ws = wb.create_sheet(title="All Registrations")
    all_ws.views.sheetView[0].showGridLines = True
    headers = ["Name", "Roll Number", "Branch", "Section", "Email", "Phone", "Club", "Registered At", "Status"]
    all_ws.append(headers)
    for cell in all_ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for r in regs:
        s = r.student
        all_ws.append([
            s.name,
            s.roll_number,
            s.branch,
            s.section or "",
            s.email,
            s.phone,
            r.club.name,
            r.registered_at.strftime("%Y-%m-%d %H:%M:%S") if r.registered_at else "",
            r.status,
        ])

    # 3. INDIVIDUAL PER-CLUB SHEETS
    regs_by_club: dict[str, list] = {}
    for r in regs:
        c_name = r.club.name[:30]
        if c_name not in regs_by_club:
            regs_by_club[c_name] = []
        regs_by_club[c_name].append(r)

    for c_name, c_regs in regs_by_club.items():
        sheet_title = "".join(c for c in c_name if c not in r"\/*?:[]")[:30]
        c_ws = wb.create_sheet(title=sheet_title)
        c_ws.views.sheetView[0].showGridLines = True
        c_headers = ["Name", "Roll Number", "Branch", "Section", "Email", "Phone", "Registered At", "Status"]
        c_ws.append(c_headers)
        for cell in c_ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for r in c_regs:
            s = r.student
            c_ws.append([
                s.name,
                s.roll_number,
                s.branch,
                s.section or "",
                s.email,
                s.phone,
                r.registered_at.strftime("%Y-%m-%d %H:%M:%S") if r.registered_at else "",
                r.status,
            ])

    # Auto-fit column widths
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"club-registrations-{club or 'all'}.xlsx"

    _log_audit(
        db,
        staff,
        action="EXPORT_GENERATED",
        details=f"Exported registrations Excel file (filter: {club or 'all'}, records: {len(regs)})",
        ip=request.client.host if request.client else None
    )

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
